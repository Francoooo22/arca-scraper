"""
export.py — Exporta comprobantes desde SQLite a Excel / CSV
Genera un archivo por CUIT o uno consolidado.
"""

import csv
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

from db import consultar_comprobantes, get_conn
from config import DB_PATH


COLUMNAS = [
    ("cuit_emisor",           "CUIT Emisor"),
    ("razon_social",          "Razón Social"),
    ("fecha_comprobante",     "Fecha"),
    ("tipo_comprobante",      "Tipo"),
    ("punto_venta",           "Pto. Venta"),
    ("numero",                "Número"),
    ("cuit_receptor",         "CUIT Receptor"),
    ("denominacion_receptor", "Denominación Receptor"),
    ("importe_neto",          "Neto"),
    ("importe_iva",           "IVA"),
    ("importe_total",         "Total"),
    ("moneda",                "Moneda"),
    ("cae",                   "CAE"),
    ("fecha_vto_cae",         "Vto. CAE"),
    ("periodo_fiscal",        "Período"),
    ("incluido_ddjj",         "En DDJJ"),
    ("observaciones",         "Observaciones"),
    ("scrapeado_en",          "Scrapeado"),
]


def exportar_excel(output_path: str = None, cuit: str = None, periodo: str = None):
    """
    Exporta a Excel con formato.
    Una hoja por CUIT si no se filtra, o una sola hoja si se filtra por CUIT.
    """
    if not EXCEL_OK:
        print("[EXPORT] openpyxl no está instalado. Usando CSV como fallback.")
        return exportar_csv(output_path, cuit, periodo)

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"comprobantes_emitidos_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja default

    # Obtener CUITs a exportar
    conn = get_conn()
    cur = conn.cursor()
    if cuit:
        cuits = [(cuit,)]
    else:
        cur.execute("SELECT DISTINCT cuit_emisor FROM comprobantes_emitidos ORDER BY cuit_emisor")
        cuits = cur.fetchall()
    conn.close()

    COLOR_HEADER = "1F3864"  # azul oscuro
    COLOR_ALT    = "EEF2F7"  # gris claro para filas alternas

    for (c,) in cuits:
        datos = consultar_comprobantes(cuit=c, periodo=periodo)
        if not datos:
            continue

        # Nombre de hoja = últimos 11 dígitos del CUIT
        razon = datos[0].get("razon_social", c)[:20] if datos else c
        nombre_hoja = f"{c[-8:]}"[:31]
        ws = wb.create_sheet(title=nombre_hoja)

        # ── Título ──────────────────────────────────────────────────────────
        ws.merge_cells(f"A1:{get_column_letter(len(COLUMNAS))}1")
        celda_titulo = ws["A1"]
        celda_titulo.value = f"Comprobantes Emitidos — {razon} ({c})"
        celda_titulo.font = Font(bold=True, color="FFFFFF", size=12)
        celda_titulo.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # ── Encabezados ─────────────────────────────────────────────────────
        for col_idx, (_, titulo) in enumerate(COLUMNAS, start=1):
            celda = ws.cell(row=2, column=col_idx, value=titulo)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            celda.alignment = Alignment(horizontal="center")

        # ── Datos ────────────────────────────────────────────────────────────
        for row_idx, comp in enumerate(datos, start=3):
            es_par = (row_idx % 2 == 0)
            fill_fila = PatternFill("solid", fgColor=COLOR_ALT) if es_par else None

            for col_idx, (campo, _) in enumerate(COLUMNAS, start=1):
                valor = comp.get(campo, "")
                celda = ws.cell(row=row_idx, column=col_idx, value=valor)
                if fill_fila:
                    celda.fill = fill_fila
                # Alinear importes a la derecha
                if campo in ("importe_neto", "importe_iva", "importe_total", "tipo_cambio"):
                    celda.alignment = Alignment(horizontal="right")
                    celda.number_format = '#,##0.00'

        # ── Anchos de columna automáticos ────────────────────────────────────
        anchos = {
            "cuit_emisor": 16, "razon_social": 30, "fecha_comprobante": 12,
            "tipo_comprobante": 16, "punto_venta": 10, "numero": 12,
            "cuit_receptor": 16, "denominacion_receptor": 30,
            "importe_neto": 14, "importe_iva": 12, "importe_total": 14,
            "moneda": 8, "cae": 16, "fecha_vto_cae": 12,
            "periodo_fiscal": 10, "incluido_ddjj": 8,
            "observaciones": 25, "scrapeado_en": 18,
        }
        for col_idx, (campo, _) in enumerate(COLUMNAS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(campo, 14)

        # ── Fila de totales ──────────────────────────────────────────────────
        ultima_fila = 2 + len(datos) + 1
        ws.cell(row=ultima_fila, column=1, value="TOTALES").font = Font(bold=True)

        campos_num = ["importe_neto", "importe_iva", "importe_total"]
        for campo in campos_num:
            col_idx = next(i+1 for i, (c, _) in enumerate(COLUMNAS) if c == campo)
            letra = get_column_letter(col_idx)
            celda = ws.cell(
                row=ultima_fila, column=col_idx,
                value=f"=SUM({letra}3:{letra}{ultima_fila-1})"
            )
            celda.font = Font(bold=True)
            celda.number_format = '#,##0.00'

    wb.save(output_path)
    print(f"[EXPORT] ✓ Excel guardado: {output_path}")
    return output_path


def exportar_csv(output_path: str = None, cuit: str = None, periodo: str = None):
    """Exporta a CSV plano (fallback o para importar en Nacional Software)."""
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"comprobantes_emitidos_{ts}.csv"

    datos = consultar_comprobantes(cuit=cuit, periodo=periodo)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        campos = [c for c, _ in COLUMNAS]
        writer = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(datos)

    print(f"[EXPORT] ✓ CSV guardado: {output_path} ({len(datos)} registros)")
    return output_path


def resumen_por_periodo(cuit: str = None) -> list[dict]:
    """Genera resumen agrupado por período, CUIT y tipo_operacion."""
    conn = get_conn()
    cur = conn.cursor()
    query = """
        SELECT
            cuit_emisor,
            razon_social,
            COALESCE(tipo_operacion, 'emitido') as tipo_operacion,
            periodo_fiscal,
            COUNT(*) as cantidad,
            SUM(importe_neto)   as total_neto,
            SUM(importe_iva)    as total_iva,
            SUM(importe_total)  as total_general,
            SUM(CASE WHEN incluido_ddjj = 1 THEN 1 ELSE 0 END) as en_ddjj
        FROM comprobantes_emitidos
    """
    params = []
    if cuit:
        query += " WHERE cuit_emisor = ?"
        params.append(cuit)
    query += """
        GROUP BY cuit_emisor, tipo_operacion, periodo_fiscal
        ORDER BY periodo_fiscal DESC, cuit_emisor, tipo_operacion
    """
    cur.execute(query, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows
